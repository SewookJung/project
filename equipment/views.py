import json
import os
import urllib
from collections import Counter

from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.conf import settings
from django.db.models import Count
from django.core import serializers

from wsgiref.util import FileWrapper
from openpyxl import load_workbook

from common.models import Client, Product, ProductModel, Mnfacture
from .models import EquipmentAttachment, Equipment, Stock, StockAttachment
from .forms import EquipmentForm
from utils.constant import REPORT_PERMISSION_DEFAULT, SAMPLE_FILE_ID, STOCK_SAMPLE_FILE_ID
from utils.constant import STATUS_KEEP, STATUS_SOLD, STATUS_RETURN, STATUS_DISPOSAL


from utils.functions import (
    make_response,
)


@login_required
def equipment_main(request):
    equipment_form = EquipmentForm()
    equipments = Equipment.objects.all()
    return render(request, 'equipment/equipment_main.html', {'equipments': equipments, 'equipment_form': equipment_form,  'permission': REPORT_PERMISSION_DEFAULT})


@login_required
def equipment_form(request):
    equipment_form = EquipmentForm()
    products = Product.objects.values('id',
                                      'name', 'makers', 'level').order_by('makers', 'level')
    return render(request, 'equipment/equipment_form.html', {"products": products, "equipment_form": equipment_form, 'permission': REPORT_PERMISSION_DEFAULT})


@login_required
def equipment_form_apply(request):
    if request.method == "POST":
        try:
            client = request.POST['client']
            product = request.POST['product_id']
            mnfacture = request.POST['mnfacture']
            product_model = request.POST['product_model']
            serial = request.POST['serial'].replace(" ", "").strip().upper()
            manager = request.POST['manager']
            location = request.POST['location']
            install_date = request.POST['install-date']
            maintenance_date = request.POST['maintenance-date']
            comments = request.POST['comments']

            try:
                stock = Stock.objects.get(serial=serial)
                raise
            except Stock.DoesNotExist:
                pass
            except:
                return make_response(status=400, content=json.dumps({'success': False, 'msg': "이미 재고에 등록 되어 있는 시리얼 번호 입니다.\n확인 후 다시 입력하시기 바랍니다.", 'error': 'serial_error'}))

            try:
                equipment = Equipment.objects.get(serial=serial)
                raise
            except Equipment.DoesNotExist:
                equipment = Equipment(client_id=client, product_id=product, product_model_id=product_model, serial=serial, mnfacture_id=mnfacture,
                                      manager=manager, location=location, install_date=install_date, maintenance_date=maintenance_date, comments=comments, creator_id=request.session['id'])
                equipment.save()
                return make_response(status=200, content=json.dumps({'success': True, 'msg': "납품 등록이 완료 되었습니다."}))
            except:
                return make_response(status=400, content=json.dumps({'success': False, 'msg': "이미 납품 된 장비 시리얼 번호 입니다.\n확인 후 다시 입력하시기 바랍니다.", 'error': 'serial_error'}))

        except:
            return make_response(status=400, content=json.dumps({'success': False, 'msg': "납품 등록에 실패 하였습니다.\n다시 시도해주시기 바랍니다."}))
    else:
        return make_response(status=400, content=json.dumps({'success': False, 'msg': "제품등록 페이지에 다시 접속해주시기 바랍니다.", 'error': 'upload_error'}))


@login_required
def equipment_detail(request, equipment_id):
    if request.method == "GET":
        equipment_form = EquipmentForm()
        products = Product.objects.values('id',
                                          'name', 'makers', 'level').order_by('makers', 'level')
        equipment = Equipment.objects.get(id=equipment_id)
        try:
            stock = Stock.objects.get(
                serial=equipment.serial, status=STATUS_SOLD)
        except:
            return render(request, 'equipment/equipment_detail.html', {"equipment": equipment, "products": products, "equipment_form": equipment_form, 'login_id': request.session['id'], 'permission': REPORT_PERMISSION_DEFAULT, })
    return render(request, 'equipment/equipment_detail.html', {"equipment": equipment, "products": products, "equipment_form": equipment_form, 'login_id': request.session['id'], 'permission': REPORT_PERMISSION_DEFAULT, 'stock': stock})


@login_required
def equipment_client_detail(request, client_id):
    if request.method == "GET":
        equipments = Equipment.objects.filter(client_id=client_id)
        client = str(equipments[0].client)
        count = equipments.count()
        return render(request, "equipment/equipment_client_detail.html", {'equipments': equipments, 'client': client, 'count': count, 'permission': REPORT_PERMISSION_DEFAULT, })


@login_required
def equipment_mnfacture_detail(request, client_id, mnfacture_id):
    if request.method == "GET":
        try:
            equipments = Equipment.objects.filter(
                client_id=client_id, mnfacture_id=mnfacture_id)
            client = equipments[0].client
            mnfacture = equipments[0].mnfacture
            count = equipments.count()
            return render(request, "equipment/equipment_mnfacture_detail.html", {'equipments': equipments, 'client': client, 'count': count, 'mnfacture': mnfacture, 'permission': REPORT_PERMISSION_DEFAULT})
        except:
            return render(request, "equipment/equipment_mnfacture_detail.html", {'permission': REPORT_PERMISSION_DEFAULT})


@login_required
def equipment_model_detail(request, client_id, mnfacture_id, model_id):
    if request.method == "GET":
        equipments = Equipment.objects.filter(
            client_id=client_id, mnfacture_id=mnfacture_id, product_model_id=model_id)
        client = equipments[0].client
        mnfacture = equipments[0].mnfacture
        product_model = equipments[0].product_model
        count = equipments.count()
        return render(request, "equipment/equipment_model_detail.html", {'equipments': equipments, 'client': client, 'count': count, 'mnfacture': mnfacture, 'product_model': product_model, 'permission': REPORT_PERMISSION_DEFAULT})


@login_required
def equipment_detail_apply(request, equipment_id):
    if request.method == "POST":
        client = request.POST['client']
        product = request.POST['product_id']
        mnfacture = request.POST['mnfacture']
        product_model = request.POST['product_model']
        manager = request.POST['manager']
        serial = request.POST['serial'].replace(" ", "").strip().upper()
        location = request.POST['location']
        install_date = request.POST['install-date']
        comments = request.POST['comments']

        try:
            stock = Stock.objects.get(serial=serial, status=STATUS_KEEP)
            raise
        except Stock.DoesNotExist:
            pass
        except:
            return make_response(status=400, content=json.dumps({'success': False, 'msg': "입력한 시리얼 번호는 이미 재고에 등록 되어 있습니다. \n시리얼 번호 확인 후 다시 시도하시기 바랍니다.", 'error': 'serial_error'}))

        try:
            equipment = Equipment.objects.get(id=int(equipment_id))

            if serial == equipment.serial:
                equipment.client_id = client
                equipment.product_id = product
                equipment.mnfactureauclf_id = mnfacture
                equipment.product_model_id = product_model
                equipment.manager = manager
                equipment.serial = serial
                equipment.location = location
                equipment.install_date = install_date
                equipment.comments = comments
                equipment.save()
                return make_response(status=200, content=json.dumps({'success': True, 'msg': "제품수정을 완료하였습니다."}))

            else:
                try:
                    exists_check_equipment = Equipment.objects.get(
                        serial=serial)

                    if exists_check_equipment.id == int(equipment_id):
                        exists_check_equipment.client_id = client
                        exists_check_equipment.product_id = product
                        exists_check_equipment.mnfactureauclf_id = mnfacture
                        exists_check_equipment.product_model_id = product_model
                        exists_check_equipment.manager = manager
                        exists_check_equipment.serial = serial
                        exists_check_equipment.location = location
                        exists_check_equipment.install_date = install_date
                        exists_check_equipment.comments = comments
                        exists_check_equipment.save()
                    else:
                        raise

                except Equipment.DoesNotExist:
                    equipment.client_id = client
                    equipment.product_id = product
                    equipment.mnfactureauclf_id = mnfacture
                    equipment.product_model_id = product_model
                    equipment.manager = manager
                    equipment.serial = serial
                    equipment.location = location
                    equipment.install_date = install_date
                    equipment.comments = comments
                    equipment.save()
                    return make_response(status=200, content=json.dumps({'success': True, 'msg': "제품수정을 완료하였습니다."}))
                except:
                    return make_response(status=400, content=json.dumps({'success': False, 'msg': "입력한 시리얼 번호는 이미 납품된 시리얼 번호입니다. \n시리얼 번호 확인 후 다시 시도하시기 바랍니다.", 'error': 'serial_error'}))

        except Equipment.DoesNotExist:
            return make_response(status=400, content=json.dumps({'success': False, 'msg': "제품정보를 가지고 오는데 실패하였습니다. \n다시 시도하시기 바랍니다.", 'error': 'upload_error'}))
        except:
            return make_response(status=400, content=json.dumps({'success': False, 'msg': "제품수정에 실패하였습니다. \n다시 시도하시기 바랍니다.", 'error': 'upload_error'}))
    else:
        return make_response(status=400, content=json.dumps({'success': False, 'msg': "제품수정에 실패하였습니다. \n다시 시도하시기 바랍니다.", 'error': 'upload_error'}))


@login_required
def equipment_upload(request):
    return render(request, 'equipment/equipment_upload.html', {'permission': REPORT_PERMISSION_DEFAULT})


@login_required
def equipment_info(request):
    client = Client.objects.all()
    mnfacture = Mnfacture.objects.all()
    product = Product.objects.all()
    product_model = ProductModel.objects.all()
    return render(request, 'equipment/equipment_info.html', {'client': client, 'mnfacture': mnfacture, 'product': product, 'product_model': product_model, 'permission': REPORT_PERMISSION_DEFAULT})


@login_required
@csrf_exempt
def equipment_upload_check(request):
    if request.method == 'POST' and request.FILES['uploadfile']:
        file = request.FILES['uploadfile']
        load_wb = load_workbook(file, data_only=True)
        line_num = 2
        err_equip_list = []
        serial_list = []

        if not os.path.exists(settings.MEDIA_ROOT):
            err_equip_list.append(
                {'no': '-', 'msg': 'NAS서버와 연결이 해제되어 파일 업로드가 불가능합니다.\n 관리자에게 문의 바랍니다.'})
            return render(request, 'equipment/equipment_upload_check.html', {'err_equip_list': err_equip_list, 'permission': REPORT_PERMISSION_DEFAULT})

        try:
            load_ws = load_wb['장비운용현황']
        except:
            err_equip_list.append(
                {'no': '-', 'msg': '시트명을 확인하세요.\n 기본 시트명은 "장비운용현황" 입니다.'})
            return render(request, 'equipment/equipment_upload_check.html', {'err_equip_list': err_equip_list, 'permission': REPORT_PERMISSION_DEFAULT})

        iter_rows = iter(load_ws.rows)
        next(iter_rows)
        max_rows = load_ws.max_row

        for row in iter_rows:
            if row[0].value == None and row[1].value == None and row[2].value == None and row[3].value == None and row[4].value == None and row[5].value == None and row[6].value == None and row[7].value == None and row[8].value == None:
                pass
            else:
                client_name = row[0].value
                mnfacture_name = row[1].value
                product_name = row[2].value
                model_name = row[3].value
                serial = row[4].value.replace(" ", "").strip().upper()
                err_dic = {'no': line_num, 'msg': ''}

                try:
                    client = Client.objects.get(name=client_name)
                except Client.DoesNotExist:
                    err_dic['msg'] = '고객사명을 확인하세요.'

                try:
                    mnfacture = Mnfacture.objects.get(
                        manafacture=mnfacture_name)
                except Mnfacture.DoesNotExist:
                    err_dic['msg'] = err_dic['msg'] + ' 제조사를 확인하세요.'

                try:
                    product = Product.objects.get(name=product_name)
                except Product.DoesNotExist:
                    err_dic['msg'] = err_dic['msg'] + ' 제품명을 확인하세요. '

                try:
                    product_model = ProductModel.objects.get(name=model_name)
                except ProductModel.DoesNotExist:
                    err_dic['msg'] = err_dic['msg'] + ' 모델명을 확인하세요.'

                try:
                    equipment = Equipment.objects.get(serial=serial)
                    raise
                except Equipment.DoesNotExist:
                    try:
                        serial_list.index(serial)
                        err_dic['msg'] = err_dic['msg'] + \
                            ' 엑셀파일에 중복되는 시리얼이 존재합니다.'
                    except ValueError:
                        serial_list.append(serial)
                except:
                    err_dic['msg'] = err_dic['msg'] + \
                        ' 이미 납품된 시리얼 정보입니다.'

                try:
                    stock = Stock.objects.get(
                        serial=serial, status=STATUS_KEEP)
                    raise
                except Stock.DoesNotExist:
                    pass
                except:
                    err_dic['msg'] = err_dic['msg'] + \
                        ' 재고에 등록 되어 있는 시리얼 정보입니다.'

                if len(err_dic['msg']) != 0:
                    err_equip_list.append(err_dic)

                line_num = line_num + 1
        if len(err_equip_list) == 0:
            equipment_attach = EquipmentAttachment(member_id=request.session['id'],
                                                   attach=file, attach_name=file.name, content_size=file.size, content_type=file.content_type)
            equipment_attach.save()
        else:
            return render(request, 'equipment/equipment_upload_check.html', {'err_equip_list': err_equip_list, 'permission': REPORT_PERMISSION_DEFAULT})

    else:
        err_equip_list.append({'no': 1, 'msg': '입력된 파일을 확인하세요'})
    return render(request, 'equipment/equipment_upload_check.html', {'err_equip_list': err_equip_list, 'equipment_attach_id': equipment_attach.id, 'permission': REPORT_PERMISSION_DEFAULT})


@login_required
@csrf_exempt
def equipment_upload_complete(request):
    client_objects = Client.objects.all()
    mnfacture_objects = Mnfacture.objects.all()
    product_objects = Product.objects.all()
    product_model_objects = ProductModel.objects.all()
    equipment_attach_id = request.POST['equipment-attach-id']

    equipment_attach = EquipmentAttachment.objects.get(id=equipment_attach_id)
    load_wb = load_workbook(equipment_attach.attach, data_only=True)
    load_ws = load_wb['장비운용현황']
    iter_rows = iter(load_ws.rows)
    next(iter_rows)

    all_values = []
    for row in iter_rows:
        cell_values = []
        for cell in row:
            cell_values.append(cell.value)
        if cell_values[0] == None:
            pass
        else:
            all_values.append(cell_values)

    for item in all_values:
        client = client_objects.values('id').get(name=item[0])
        mnfacture = mnfacture_objects.values('id').get(manafacture=item[1])
        product = product_objects.values('id').get(name=item[2])
        product_model = product_model_objects.values('id').get(name=item[3])
        serial = item[4].replace(" ", "").strip().upper()
        location = item[5]
        manager = item[6]
        install_date = item[7]
        comments = item[8]

        if manager == None:
            manager = ""

        if comments == None:
            comments = ""
        else:
            comments = item[8]

        equipment = Equipment(client_id=client['id'], product_id=product['id'], product_model_id=product_model['id'], mnfacture_id=mnfacture['id'],
                              serial=serial, location=location, manager=manager, install_date=install_date, comments=comments, creator_id=request.session['id'])
        equipment.save()

    return redirect("equipment_main")


@login_required
def equipment_upload_cancel(request):
    if request.method == 'POST':
        equipment_attach_id = request.POST['equipment_attach_id']
        try:
            EquipmentAttachment.objects.get(id=equipment_attach_id).delete()
            return make_response(status=200, content=json.dumps({'success': True, 'msg': "장비현황 등록을 취소하였습니다."}))
        except:
            return make_response(status=400, content=json.dumps({'success': False, 'error': "장비현황 등록을 취소하는데 실패하였습니다. \n 파일을 다시 업로드 해주시길 바랍니다."}))
    else:
        return make_response(status=400, content=json.dumps({'success': False, 'error': "장비현황 등록을 취소하는데 실패하였습니다. \n 파일을 다시 업로드 해주시길 바랍니다."}))


@login_required
def equipment_all_list(request, client_id):
    if request.method == 'GET':
        equipments = Equipment.objects.filter(
            client=client_id).order_by('mnfacture')

        if equipments.count() == 0:
            return make_response(status=400, content=json.dumps({'success': False, 'error': "❌ 선택하신 고객사의 납품 현황이 존재하지 않습니다.\n      고객사를 다시 선택해주시기 바랍니다."}))
        else:
            equipment_lists = {}
            for equipment in equipments:
                client = str(equipment.client)
                mnfacture = str(equipment.mnfacture)
                product_model = str(equipment.product_model)

                client_list = list(equipment_lists.keys())
                if not client in client_list:
                    equipment_lists[client] = {}
                    equipment_lists['totalCount'] = equipments.count()

                mnfacture_list = list(equipment_lists[client].keys())
                if not mnfacture in mnfacture_list:
                    equipment_lists[client][mnfacture] = {}

                product_model_list = list(
                    equipment_lists[client][mnfacture].keys())
                if not product_model in product_model_list:
                    equipment_lists[client][mnfacture][product_model] = {
                        'count': 1}
                else:
                    equipment_lists[client][mnfacture][product_model]['count'] += 1
            return make_response(status=200, content=json.dumps({'success': True, 'equipment_lists': equipment_lists}))


@login_required
def equipment_client_mnfacture_list(request, client_id):
    if request.method == 'GET':
        equipments = Equipment.objects.filter(client=client_id)
        mnfactures = []

        for equipment in equipments:
            mnfacture_id = equipment.mnfacture_id
            mnfacture_name = str(equipment.mnfacture)
            data = {'mnfacture_id': mnfacture_id,
                    'mnfacture_name': mnfacture_name}
            if not data in mnfactures:
                mnfactures.append(data)

        return make_response(status=200, content=json.dumps({'success': True, 'mnfactures': mnfactures}))


@login_required
def equipment_selected_list(request, client_id, mnfacture_id):
    equipments = Equipment.objects.filter(
        client=client_id, mnfacture=mnfacture_id).order_by('mnfacture')
    client_equipments_total_count = Equipment.objects.filter(
        client=client_id).count()
    equipment_lists = {}
    for equipment in equipments:
        client = str(equipment.client)
        mnfacture = str(equipment.mnfacture)
        product_model = str(equipment.product_model)

        client_list = list(equipment_lists.keys())
        if not client in client_list:
            equipment_lists[client] = {}
            equipment_lists['totalCount'] = client_equipments_total_count

        mnfacture_list = list(equipment_lists[client].keys())
        if not mnfacture in mnfacture_list:
            equipment_lists[client][mnfacture] = {}

        product_model_list = list(
            equipment_lists[client][mnfacture].keys())
        if not product_model in product_model_list:
            equipment_lists[client][mnfacture][product_model] = {
                'count': 1}
        else:
            equipment_lists[client][mnfacture][product_model]['count'] += 1

    return make_response(status=200, content=json.dumps({'success': True, 'equipment_lists': equipment_lists}))


@login_required
def equipment_delete(request, equipment_id):
    if request.method == 'GET':
        try:
            equipment = Equipment.objects.get(id=equipment_id)
            equipment.delete()

            try:
                serial = equipment.serial
                stock = Stock.objects.get(serial=serial)
                stock.status = STATUS_KEEP
                stock.save()
                return make_response(status=200, content=json.dumps({'success': True, 'msg': "해당 제품정보를 삭제하였습니다."}))

            except:
                return make_response(status=200, content=json.dumps({'success': True, 'msg': "해당 제품정보를 삭제하였습니다."}))

            return make_response(status=200, content=json.dumps({'success': True, 'msg': "해당 제품정보를 삭제하였습니다."}))
        except:
            return make_response(status=400, content=json.dumps({'success': True, 'msg': "해당 제품정보를 삭제하는데 실패 하였습니다."}))
    else:
        return make_response(status=400, content=json.dumps({'success': True, 'msg': "해당 제품정보를 삭제하는데 실패 하였습니다."}))


@login_required
def equipment_download_check(request):
    if not os.path.exists(settings.MEDIA_ROOT):
        return make_response(status=400, content=json.dumps({'success': False, 'msg': "NAS서버와 연결이 해제되어 파일 다운로드가 불가능합니다.\n관리자에게 문의 바랍니다."}))
    try:
        sample = EquipmentAttachment.objects.get(id=SAMPLE_FILE_ID)
        return make_response(status=200, content=json.dumps({'success': True}))
    except EquipmentAttachment.DoesNotExist:
        return make_response(status=400, content=json.dumps({'success': False, 'msg': "파일이 존재하지 않아 다운로드 할 수 없습니다.\n관리자에게 문의 바랍니다."}))


@login_required
def equipment_sample_download(request):
    attach_info = EquipmentAttachment.objects.get(id=SAMPLE_FILE_ID)
    filename = os.path.join(settings.MEDIA_ROOT, attach_info.attach.name)
    response = ""
    if os.path.exists(filename):
        with open(filename, 'rb') as f:
            response = HttpResponse(FileWrapper(
                f), content_type=attach_info.attach)
            response['Content-Disposition'] = 'attachment; filename*=UTF-8\'\'%s' % urllib.parse.quote(
                attach_info.attach_name.encode('utf-8'))
    return response


@login_required
def equipment_stock(request):
    equipment_form = EquipmentForm()
    return render(request, 'equipment/equipment_stock.html', {"equipment_form": equipment_form, 'permission': REPORT_PERMISSION_DEFAULT})


@login_required
def equipment_stock_get_list(request):
    try:
        mnfacture_id = request.POST['mnfactureId']

        stock_objects = Stock.objects.filter(
            mnfacture_id=mnfacture_id).order_by("product_model")
        mn_objects = Mnfacture.objects.all()
        product_model_objects = ProductModel.objects.all()
        stocks = []

        mnfacture_name = str(mn_objects.get(id=mnfacture_id))
        mnfacture_object = {mnfacture_name: []}
        stocks.append(mnfacture_object)
        product_model_id_list = stock_objects.values(
            "product_model").distinct()

        for product in product_model_id_list:
            product_model_id = product['product_model']

            product_model_name = str(
                product_model_objects.get(id=product_model_id))
            stock_keep_count = stock_objects.filter(
                product_model_id=product_model_id, status=STATUS_KEEP).count()
            stock_sold_count = stock_objects.filter(
                product_model_id=product_model_id, status=STATUS_SOLD).count()
            stock_disposal_count = stock_objects.filter(
                product_model_id=product_model_id, status=STATUS_DISPOSAL).count()
            stock_return_count = stock_objects.filter(
                product_model_id=product_model_id, status=STATUS_RETURN).count()
            stock_object = {product_model_name: {
                'id': product_model_id, 'keep': stock_keep_count, 'sold': stock_sold_count, 'disposal': stock_disposal_count, 'return': stock_return_count, 'total': stock_keep_count + stock_sold_count}}
            mnfacture_object[mnfacture_name].append(stock_object)
        return make_response(status=200, content=json.dumps({'success': True, 'stocks': stocks}))

    except:
        return make_response(status=400, content=json.dumps({'success': False, 'msg': "재고현황을 가지고 있는데 실패 하였습니다.\n관리자에게 문의 바랍니다."}))


@login_required
def equipment_stock_check(request):
    if not os.path.exists(settings.MEDIA_ROOT):
        return make_response(status=400, content=json.dumps({'success': False, 'msg': "NAS서버와 연결이 해제되어 파일 다운로드가 불가능합니다.\n관리자에게 문의 바랍니다."}))
    try:
        sample = EquipmentAttachment.objects.get(id=SAMPLE_FILE_ID)
        return make_response(status=200, content=json.dumps({'success': True}))
    except EquipmentAttachment.DoesNotExist:
        return make_response(status=400, content=json.dumps({'success': False, 'msg': "파일이 존재하지 않아 다운로드 할 수 없습니다.\n관리자에게 문의 바랍니다."}))


@login_required
def equipment_stock_form(request):
    equipment_form = EquipmentForm()
    products = Product.objects.values('id',
                                      'name', 'makers', 'level').order_by('makers', 'level')
    return render(request, 'equipment/equipment_stock_form.html', {"products": products, "equipment_form": equipment_form, 'permission': REPORT_PERMISSION_DEFAULT})


@login_required
def equipment_stock_form_apply(request):
    if request.method == "POST":
        equipments = Equipment.objects.all()
        stocks = Stock.objects.all()

        try:
            mnfacture = request.POST['mnfacture']
            product_model = request.POST['product_model']
            serial = request.POST['serial'].replace(" ", "").strip().upper()
            location = request.POST['location']
            receive_date = request.POST['install-date']
            comments = request.POST['comments']

            try:
                equipment = Equipment.objects.get(serial=serial)
                raise
            except Equipment.DoesNotExist:
                pass
            except:
                return make_response(status=400, content=json.dumps({'success': False, 'msg': "납품 된 장비 Serial 입니다.\n확인 후 다시 입력하시기 바랍니다.", 'error': 'serial_error'}))

            try:
                stock = Stock.objects.get(serial=serial)
                raise
            except Stock.DoesNotExist:
                new_stock = Stock(mnfacture_id=mnfacture, product_model_id=product_model, serial=serial,
                                  location=location, status=STATUS_KEEP, receive_date=receive_date, creator_id=request.session['id'], comments=comments)
                new_stock.save()
                return make_response(status=200, content=json.dumps({'success': True, 'msg': "재고 등록이 완료 되었습니다."}))
            except:
                return make_response(status=400, content=json.dumps({'success': False, 'msg': "재고에 이미 등록되어 있는 Serial 입니다.\n확인 후 다시 입력하시기 바랍니다.", 'error': 'serial_error'}))

        except:
            return make_response(status=400, content=json.dumps({'success': False, 'msg': "재고 등록에 실패 하였습니다.\n다시 시도해주시기 바랍니다."}))
    else:
        return make_response(status=400, content=json.dumps({'success': False, 'msg': "제품등록 페이지에 다시 접속해주시기 바랍니다."}))


@login_required
def equipment_stock_multi_apply(request):
    if request.method == 'POST':
        stocks = Stock.objects.all()
        client_id = request.POST['client']
        stock_ids = request.POST.getlist('stockIds[]')
        location = request.POST['location']
        delivery_date = request.POST['deliveryDate']
        maintenance_date = request.POST['maintenanceDate']

        try:
            for stock_id in stock_ids:
                stock = stocks.get(id=stock_id)
                stock.status = STATUS_SOLD
                stock.save()

                equipment = Equipment(client_id=client_id, product_model_id=stock.product_model.id, serial=stock.serial, mnfacture_id=stock.mnfacture.id,
                                      location=location, install_date=delivery_date, maintenance_date=maintenance_date, comments="", creator_id=request.session['id'])
                equipment.save()
            return make_response(status=200, content=json.dumps({'success': True, 'msg': "납품이 완료 되었습니다."}))
        except:
            return make_response(status=400, content=json.dumps({'success': True, 'msg': "납품이 정상적으로 처리 되지 않았습니다.\n다시 시도해 주세요."}))


@login_required
def equipment_stock_detail(request, model_id, model_status):
    stock_objects = Stock.objects.all()
    stocks = stock_objects.filter(
        product_model_id=model_id, status=model_status)
    stock_return_count = stock_objects.filter(
        product_model_id=model_id, status=STATUS_RETURN).count()
    stock_sold_count = stock_objects.filter(
        product_model_id=model_id, status=STATUS_SOLD).count()
    stock_keep_count = stock_objects.filter(
        product_model_id=model_id, status=STATUS_KEEP).count()
    stock_disposal_count = stock_objects.filter(
        product_model_id=model_id, status=STATUS_DISPOSAL).count()
    stock_model_name = stocks[0].product_model
    stock_mnfacture_name = stocks[0].mnfacture
    stock_product_model = stocks[0].product_model_id
    total_count = stocks.count()
    equipment_form = EquipmentForm()

    if model_status == STATUS_KEEP:
        return render(request, 'equipment/equipment_stock_keep.html', {'stocks': stocks, 'stock_model_name': stock_model_name, 'stock_mnfacture_name': stock_mnfacture_name, 'equipment_form': equipment_form, 'permission': REPORT_PERMISSION_DEFAULT, 'stock_return_count': stock_return_count, "stock_sold_count": stock_sold_count, "stock_disposal_count": stock_disposal_count, 'stock_product_model': stock_product_model, 'total_count': total_count})

    if model_status == STATUS_SOLD:
        equipments = Equipment.objects.all()
        stock_infos = stocks.values('serial', 'receive_date')

        solds = []
        for stock_info in stock_infos:
            equipment_data = equipments.get(serial=stock_info['serial'])
            equipment_data.receive_date = stock_info['receive_date']
            solds.append(equipment_data)
        return render(request, 'equipment/equipment_stock_sold.html', {'solds': solds, 'stock_model_name': stock_model_name, 'stock_mnfacture_name': stock_mnfacture_name, 'equipment_form': equipment_form, 'permission': REPORT_PERMISSION_DEFAULT, 'stock_keep_count': stock_keep_count, "stock_return_count": stock_return_count, "stock_disposal_count": stock_disposal_count, 'stock_product_model': stock_product_model, 'total_count': total_count})

    if model_status == STATUS_RETURN:
        return render(request, 'equipment/equipment_stock_return.html', {'returns': stocks, 'stock_model_name': stock_model_name, 'stock_mnfacture_name': stock_mnfacture_name, 'equipment_form': equipment_form, 'permission': REPORT_PERMISSION_DEFAULT, 'stock_keep_count': stock_keep_count, "stock_sold_count": stock_sold_count, "stock_disposal_count": stock_disposal_count, 'stock_product_model': stock_product_model,  'total_count': total_count})

    if model_status == STATUS_DISPOSAL:
        return render(request, 'equipment/equipment_stock_disposal.html', {'disposals': stocks, 'stock_model_name': stock_model_name, 'stock_mnfacture_name': stock_mnfacture_name, 'equipment_form': equipment_form, 'permission': REPORT_PERMISSION_DEFAULT, 'stock_keep_count': stock_keep_count, "stock_return_count": stock_return_count, "stock_sold_count": stock_sold_count, 'stock_product_model': stock_product_model,  'total_count': total_count})


@login_required
def equipment_stock_edit(request, stock_id):
    if request.method == "GET":
        equipment_form = EquipmentForm()
        products = Product.objects.values('id',
                                          'name', 'makers', 'level').order_by('makers', 'level')
        stock = Stock.objects.get(id=stock_id)
    return render(request, 'equipment/equipment_stock_edit.html', {"stock": stock, "products": products, "equipment_form": equipment_form, 'login_id': request.session['id'], 'permission': REPORT_PERMISSION_DEFAULT})


@login_required
def equipment_stock_detail_apply(request, stock_id):
    if request.method == "POST":
        product = request.POST['product_id']
        mnfacture = request.POST['mnfacture']
        product_model = request.POST['product_model']
        serial = request.POST['serial'].replace(" ", "").strip().upper()
        location = request.POST['location']
        receive_date = request.POST['install-date']
        comments = request.POST['comments']

        try:
            equipment = Equipment.objects.get(serial=serial)
            raise
        except Equipment.DoesNotExist:
            pass
        except:
            return make_response(status=400, content=json.dumps({'success': False, 'msg': "입력한 시리얼 번호는 이미 납품되어 있는 시리얼 번호 입니다. \n시리얼번호 확인 후 다시 시도하시기 바랍니다.", 'error': 'serial_error'}))

        try:
            stock = Stock.objects.get(id=stock_id)
            if serial == stock.serial:
                stock.product_id = product
                stock.mnfacture_id = mnfacture
                stock.product_model_id = product_model
                stock.serial = serial
                stock.location = location
                stock.receive_date = receive_date
                stock.comments = comments
                stock.save()
                return make_response(status=200, content=json.dumps({'success': True, 'msg': "제품수정을 완료하였습니다."}))

            else:
                try:
                    exists_check_stock = Stock.objects.get(
                        serial=serial)

                    if exists_check_stock.id == stock_id:
                        exists_check_stock.product_id = product
                        exists_check_stock.mnfactureauclf_id = mnfacture
                        exists_check_stock.product_model_id = product_model
                        exists_check_stock.serial = serial
                        exists_check_stock.location = location
                        exists_check_stock.receive_date = receive_date
                        exists_check_stock.comments = comments
                        exists_check_stock.save()
                    else:
                        raise

                except Stock.DoesNotExist:
                    stock.product_id = product
                    stock.mnfactureauclf_id = mnfacture
                    stock.product_model_id = product_model
                    stock.serial = serial
                    stock.location = location
                    stock.receive_date = receive_date
                    stock.comments = comments
                    stock.save()
                    return make_response(status=200, content=json.dumps({'success': True, 'msg': "제품수정을 완료하였습니다."}))
                except:
                    return make_response(status=400, content=json.dumps({'success': False, 'msg': "입력한 시리얼 번호가 재고에 이미 존재합니다. \n시리얼번호 확인 후 다시 시도하시기 바랍니다.", 'error': 'serial_error'}))

        except Equipment.DoesNotExist:
            return make_response(status=400, content=json.dumps({'success': False, 'msg': "제품정보를 가지고 오는데 실패하였습니다. \n다시 시도하시기 바랍니다.", 'error': 'upload_error'}))
        except:
            return make_response(status=400, content=json.dumps({'success': False, 'msg': "제품수정에 실패하였습니다. \n다시 시도하시기 바랍니다.", 'error': 'upload_error'}))
    else:
        return make_response(status=400, content=json.dumps({'success': False, 'msg': "제품수정에 실패하였습니다. \n다시 시도하시기 바랍니다.", 'error': 'upload_error'}))


@login_required
def equipment_stock_delete(request, stock_id):
    if request.method == 'GET':
        try:
            Stock.objects.get(id=stock_id).delete()
            return make_response(status=200, content=json.dumps({'success': True, 'msg': "해당 제품정보를 삭제하였습니다."}))
        except:
            return make_response(status=400, content=json.dumps({'success': True, 'msg': "해당 제품정보를 삭제하는데 실패 하였습니다."}))
    else:
        return make_response(status=400, content=json.dumps({'success': True, 'msg': "해당 제품정보를 삭제하는데 실패 하였습니다."}))


def equipment_stock_upload(request):
    return render(request, 'equipment/equipment_stock_upload.html', {'permission': REPORT_PERMISSION_DEFAULT})


@login_required
def equipment_stock_delivery_apply(request):
    if request.method == 'POST':
        stocks = Stock.objects.all()
        client = request.POST['client']
        stock_id = request.POST['stockId']
        location = request.POST['location']
        delivery_date = request.POST['deliveryDate']
        maintenance_date = request.POST['maintenanceDate']
        try:
            stock = stocks.get(id=stock_id)
            stock.status = "sold"
            stock.save()

            equipment = Equipment(client_id=client, product_model_id=stock.product_model.id, serial=stock.serial, maintenance_date=maintenance_date,
                                  mnfacture_id=stock.mnfacture.id, location=location, install_date=delivery_date, comments="", creator_id=request.session['id'])
            equipment.save()
            return make_response(status=200, content=json.dumps({'success': True, 'msg': "납품이 완료 되었습니다."}))
        except:
            return make_response(status=400, content=json.dumps({'success': True, 'msg': "납품이 정상적으로 처리 되지 않았습니다.\n다시 시도해 주세요."}))


@login_required
def equipment_stock_return_apply(request):
    if request.method == "POST":
        try:
            stock_id = request.POST['returnStockId']
            return_date = request.POST['returnDate']
            return_comments = request.POST['returnComments']

            stock = Stock.objects.get(id=stock_id)
            stock.status = STATUS_RETURN
            stock.return_date = return_date
            stock.comments = return_comments
            stock.save()
            return make_response(status=200, content=json.dumps({'success': True, 'msg': "✔ 해당 재고가 정상적으로 반납 처리 되었습니다."}))
        except:
            return make_response(status=400, content=json.dumps({'success': False, 'msg': "❌ 해당 재고를 반납 처리하는데 실패하였습니다.\n      다시 시도해주세요"}))
    else:
        return make_response(status=400, content=json.dumps({'success': False, 'msg': "❌ 정상적인 접근방식이 아닙니다.\n      다시 시도해주세요"}))


@login_required
def equipment_stock_disposal_apply(request):
    if request.method == "POST":
        try:
            stock_id = request.POST['disposalStockId']
            disposal_date = request.POST['disposalDate']
            disposal_comments = request.POST['disposalComments']

            stock = Stock.objects.get(id=stock_id)
            stock.status = STATUS_DISPOSAL
            stock.disposal_date = disposal_date
            stock.comments = disposal_comments
            stock.save()
            return make_response(status=200, content=json.dumps({'success': True, 'msg': "✔ 해당 재고가 정상적으로 폐기 처리 되었습니다."}))
        except:
            return make_response(status=400, content=json.dumps({'success': False, 'msg': "❌ 해당 재고를 폐기 처리하는데 실패하였습니다.\n      다시 시도해주세요"}))
    else:
        return make_response(status=400, content=json.dumps({'success': False, 'msg': "❌ 정상적인 접근방식이 아닙니다.\n      다시 시도해주세요"}))


@login_required
@csrf_exempt
def equipment_stock_upload_check(request):
    if request.method == 'POST' and request.FILES['uploadfile']:
        file = request.FILES['uploadfile']
        load_wb = load_workbook(file, data_only=True)
        line_num = 2
        err_stock_list = []
        serial_list = []

        if not os.path.exists(settings.MEDIA_ROOT):
            err_stock_list.append(
                {'no': '-', 'msg': 'NAS서버와 연결이 해제되어 파일 업로드가 불가능합니다.\n 관리자에게 문의 바랍니다.'})
            return render(request, 'equipment/equipment_stock_upload_check.html', {'err_stock_list': err_stock_list, 'permission': REPORT_PERMISSION_DEFAULT})

        try:
            load_ws = load_wb['재고신규등록']

        except:
            err_stock_list.append(
                {'no': '-', 'msg': '시트명을 확인하세요.\n 기본 시트명은 "재고신규등록" 입니다.'})
            return render(request, 'equipment/equipment_stock_upload_check.html', {'err_stock_list': err_stock_list, 'permission': REPORT_PERMISSION_DEFAULT})

        iter_rows = iter(load_ws.rows)
        next(iter_rows)
        max_rows = load_ws.max_row

        for row in iter_rows:
            if row[0].value == None and row[1].value == None and row[2].value == None and row[3].value == None and row[4].value == None and row[5].value == None:
                pass
            else:
                mnfacture_name = row[0].value
                model_name = row[1].value
                serial = row[2].value.replace(" ", "").strip().upper()
                err_dic = {'no': line_num, 'msg': ''}

                try:
                    mnfacture = Mnfacture.objects.get(
                        manafacture=mnfacture_name)
                except Mnfacture.DoesNotExist:
                    err_dic['msg'] = err_dic['msg'] + ' 제조사를 확인하세요.'

                try:
                    product_model = ProductModel.objects.get(name=model_name)
                except ProductModel.DoesNotExist:
                    err_dic['msg'] = err_dic['msg'] + ' 모델명을 확인하세요.'

                try:
                    stock = Stock.objects.get(
                        serial=serial, status=STATUS_KEEP)
                    raise
                except Stock.DoesNotExist:
                    try:
                        serial_list.index(serial)
                        err_dic['msg'] = err_dic['msg'] + \
                            ' 엑셀파일에 중복되는 시리얼이 존재합니다.'
                    except ValueError:
                        serial_list.append(serial)
                except:
                    err_dic['msg'] = err_dic['msg'] + \
                        ' 이미 재고에 등록되어 있는 시리얼 번호입니다.'

                try:
                    equipment = Equipment.objects.get(serial=serial)
                    raise
                except Equipment.DoesNotExist:
                    pass
                except:
                    err_dic['msg'] = err_dic['msg'] + \
                        ' 이미 납품된 시리얼 정보입니다.'

                if len(err_dic['msg']) != 0:
                    err_stock_list.append(err_dic)

                line_num = line_num + 1

        if len(err_stock_list) == 0:
            stock_attach = StockAttachment(member_id=request.session['id'],
                                           attach=file, attach_name=file.name, content_size=file.size, content_type=file.content_type)
            stock_attach.save()
        else:
            return render(request, 'equipment/equipment_stock_upload_check.html', {'err_stock_list': err_stock_list, 'permission': REPORT_PERMISSION_DEFAULT})
    else:
        err_stock_list.append({'no': 1, 'msg': '입력된 파일을 확인하세요'})
    return render(request, 'equipment/equipment_stock_upload_check.html', {'err_stock_list': err_stock_list, 'stock_attach_id': stock_attach.id, 'permission': REPORT_PERMISSION_DEFAULT})


@login_required
@csrf_exempt
def equipment_stock_upload_complete(request):
    mnfacture_objects = Mnfacture.objects.all()
    product_model_objects = ProductModel.objects.all()
    stock_attach_id = request.POST['equipment-attach-id']

    equipment_attach = StockAttachment.objects.get(id=stock_attach_id)
    load_wb = load_workbook(equipment_attach.attach, data_only=True)
    load_ws = load_wb['재고신규등록']
    iter_rows = iter(load_ws.rows)
    next(iter_rows)

    all_values = []
    for row in iter_rows:
        cell_values = []
        for cell in row:
            cell_values.append(cell.value)
        if cell_values[0] == None:
            pass
        else:
            all_values.append(cell_values)

    for item in all_values:
        mnfacture = mnfacture_objects.values('id').get(manafacture=item[0])
        product_model = product_model_objects.values('id').get(name=item[1])
        serial = item[2].replace(" ", "").strip().upper()
        location = item[3]
        receive_date = item[4]

        if item[5] == None:
            comments = ""
        else:
            comments = item[5]

        stock = Stock(mnfacture_id=mnfacture['id'], product_model_id=product_model['id'], serial=serial,
                      location=location, status=STATUS_KEEP, receive_date=receive_date, creator_id=request.session['id'], comments=comments)
        stock.save()
    return redirect("equipment_stock")


@login_required
def equipment_stock_upload_cancel(request):
    if request.method == 'POST':
        equipment_attach_id = request.POST['equipment_attach_id']
        try:
            StockAttachment.objects.get(id=equipment_attach_id).delete()
            return make_response(status=200, content=json.dumps({'success': True, 'msg': "장비현황 등록을 취소하였습니다."}))
        except:
            return make_response(status=400, content=json.dumps({'success': False, 'error': "장비현황 등록을 취소하는데 실패하였습니다. \n 파일을 다시 업로드 해주시길 바랍니다."}))
    else:
        return make_response(status=400, content=json.dumps({'success': False, 'error': "장비현황 등록을 취소하는데 실패하였습니다. \n 파일을 다시 업로드 해주시길 바랍니다."}))


@login_required
def equipment_stock_download_check(request):
    if not os.path.exists(settings.MEDIA_ROOT):
        return make_response(status=400, content=json.dumps({'success': False, 'msg': "NAS서버와 연결이 해제되어 파일 다운로드가 불가능합니다.\n관리자에게 문의 바랍니다."}))
    try:
        sample = StockAttachment.objects.get(id=STOCK_SAMPLE_FILE_ID)
        return make_response(status=200, content=json.dumps({'success': True}))
    except EquipmentAttachment.DoesNotExist:
        return make_response(status=400, content=json.dumps({'success': False, 'msg': "파일이 존재하지 않아 다운로드 할 수 없습니다.\n관리자에게 문의 바랍니다."}))


@login_required
def equipment_stock_sample_download(request):
    attach_info = StockAttachment.objects.get(id=STOCK_SAMPLE_FILE_ID)
    filename = os.path.join(settings.MEDIA_ROOT, attach_info.attach.name)
    response = ""
    if os.path.exists(filename):
        with open(filename, 'rb') as f:
            response = HttpResponse(FileWrapper(
                f), content_type=attach_info.attach)
            response['Content-Disposition'] = 'attachment; filename*=UTF-8\'\'%s' % urllib.parse.quote(
                attach_info.attach_name.encode('utf-8'))
    return response


@login_required
def equipment_stock_permission_check(request):
    stocks = Stock.objects.all()
    stock_ids = request.POST.getlist('stockIds[]')
    permission_denied_lists = []
    try:
        for stock_id in stock_ids:
            creator = stocks.values("creator").get(id=stock_id)
            if not creator['creator'] == request.session['id']:
                raise
        return make_response(status=200, content=json.dumps({'success': True}))

    except:
        for stock_id in stock_ids:
            stock = stocks.get(id=stock_id)
            if not stock.creator_id == request.session['id']:
                mnfacture = str(stock.mnfacture)
                product_model = str(stock.product_model)
                creator = str(stock.creator)
                permission_denied_stock = {
                    "mnfacture": mnfacture, "product_model": product_model, "serial": stock.serial, "creator": creator}
                permission_denied_lists.append(permission_denied_stock)
        return make_response(status=400, content=json.dumps({'success': False, 'error': "permission-error", "denied_lists": permission_denied_lists}))


@login_required
def equipment_stock_multi_delete(request):
    stocks = Stock.objects.all()
    stock_ids = request.POST.getlist('stockIds[]')
    try:
        for stock_id in stock_ids:
            stocks.get(id=stock_id).delete()
        return make_response(status=200, content=json.dumps({'success': True, 'msg': "재고 삭제에 성공 하였습니다."}))
    except:
        return make_response(status=400, content=json.dumps({'success': False, 'msg': "재고 삭제에 실패하였습니다.\n다시 시도해주시기 바랍니다."}))


@login_required
def equipment_stock_multi_return(request):
    if request.method == "POST":
        try:
            stocks = Stock.objects.all()
            stock_ids = request.POST.getlist('stockIds[]')
            return_date = request.POST['returnDate']
            return_comments = request.POST['returnComments']
            for stock_id in stock_ids:
                stock = stocks.get(id=stock_id)
                stock.status = STATUS_RETURN
                stock.return_date = return_date
                stock.comments = return_comments
                stock.save()
            return make_response(status=200, content=json.dumps({'success': True, 'msg': "✔ 해당 재고가 정상적으로 반납 처리 되었습니다."}))
        except:
            return make_response(status=400, content=json.dumps({'success': False, 'msg': "❌ 해당 재고를 반납 처리하는데 실패하였습니다.\n      다시 시도해주세요"}))


@login_required
def equipment_stock_multi_disposal(request):
    if request.method == "POST":
        try:
            stocks = Stock.objects.all()
            stock_ids = request.POST.getlist('stockIds[]')
            disposal_date = request.POST['disposalDate']
            disposal_comments = request.POST['disposalComments']
            for stock_id in stock_ids:
                stock = stocks.get(id=stock_id)
                stock.status = STATUS_DISPOSAL
                stock.disposal_date = disposal_date
                stock.comments = disposal_comments
                stock.save()
            return make_response(status=200, content=json.dumps({'success': True, 'msg': "✔ 해당 재고가 정상적으로 폐기 처리 되었습니다."}))
        except:
            return make_response(status=400, content=json.dumps({'success': False, 'msg': "❌ 해당 재고를 폐기 처리하는데 실패하였습니다.\n      다시 시도해주세요"}))


@login_required
def equipment_test(request):
    equipment = EquipmentForm()
    return render(request, 'equipment/equipment_test.html', {'equipment': equipment})
